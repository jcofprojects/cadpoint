from pyautocad import Autocad, APoint
from openpyxl import load_workbook

class Coordenadas:
    def __init__(self, nome, pX, pY, pZ):
        self.nome = nome
        self.pX = pX
        self.pY = pY
        self.pZ = pZ

#Conexão com Excel
wb = load_workbook(filename = 'Coordenadas.xlsx')
sheet_ranges = wb['Planilha1']
ws = wb.active   
print("CadPoint 1.0 - By Julio Felipe") 
row_count = ws.max_row
for i in range(2, row_count - 1):
    cP1 = Coordenadas(sheet_ranges['A' + str(i)].value, sheet_ranges['B' + str(i)].value, sheet_ranges['C' + str(i)].value, sheet_ranges['D' + str(i)].value)
    print('{0}, {1}, {2}, {3}'.format(cP1.nome, cP1.pX, cP1.pY, cP1.pZ))
#------------------------------------------------
#Conexão com AutoCAD
    acad = Autocad()
    p1 = APoint(cP1.pX, cP1.pY, cP1.pZ)
    acad.model.AddPoint(p1)
    p1 = APoint(cP1.pX + 2.5, cP1.pY + 2.5, cP1.pZ)
    acad.model.AddText(cP1.nome, p1, 2.5)
    
print("Locacao de Pontos Finalizada!")